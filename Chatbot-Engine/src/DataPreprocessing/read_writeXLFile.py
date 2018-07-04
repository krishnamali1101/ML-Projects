import openpyxl
from openpyxl import Workbook


# ====================================================================
# returns dict {col1:[col2,col3...]}
def read_xl(filename):
    dict_obj = {}
    try:
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active

        for row in sheet.iter_rows():
            intent = []
            first_cell = True
            for cell in row:
                if cell.value:
                    # read first col as key
                    if first_cell:
                        first_cell = False
                        dict_obj[cell.value] = intent
                    else:
                        intent.append(cell.value)
    except:
        pass

    return dict_obj


def write_xl(filename, mydict):
    wb = Workbook()
    sheet = wb.active
    r = 1
    for key in mydict:
        # write key
        sheet.cell(row=r, column=1).value = key

        # write intents
        c = 2
        for val in mydict[key]:
            sheet.cell(row=r, column=c).value = val
            c += 1
        r += 1

    try:
        wb.save(filename)
    except OSError:
        print('\nModel File is still open...')



def test():
    myDictionary = {
        'car seat': ['$10', '12'],
        'piano': ['$12', '2'],
        'picture frame': ['$2', '14', 'IT1'],
        'shoes': ['$20', '13', 'IT1'],
    }

    filename = "example.xlsx"
    write_xl(filename, myDictionary)
    dct = read_xl(filename)

    print(dct)

# test()
