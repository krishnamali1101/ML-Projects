import csv
import openpyxl
import os
import sys
import xlrd
import xlwt
from xlsxwriter.workbook import Workbook

# using openpyxl
def read_sampledata_file(filename, sheet_name= "Input Data"):
    map_topic_description = {}
    wb = openpyxl.load_workbook(filename)
    sheet = wb.get_sheet_by_name(sheet_name)

    for i in range(9, 33, 1):
        # print(sheet.cell(row=i, column=1).value)
        # print(sheet.cell(row=i, column=2).value)
        # print("\n")
        map_topic_description[sheet.cell(row=i, column=1).value.lower()] = sheet.cell(row=i, column=2).value.lower()
    return map_topic_description


# =======================================================
total_lines =0
total_missed_lines = 0

# read csv/xls/xlsx file from dict object
# TODO modify it to read just csv and do merge cell opration in texonomy.py file
def csv2dict(filename):
    map_topic_description = {}
    with open(filename, 'r') as fp:
        # skipping field names through first row
        line = " "
        try:
            # skipping field names through first row
            line = fp.readline()
        except:
            pass

        line_number = 1
        missed_lines = 0
        global total_lines
        global total_missed_lines
        while len(line) > 0:
            try:
                line_number += 1
                line = fp.readline()
                lst = line.split("\t")
                # creating map of (topic: definition+breakdown)
                if len(lst)==2:
                    map_topic_description[lst[0]] = lst[1]
                elif len(lst)==4:
                    # special case for mortgage data
                    map_topic_description[lst[1]] = lst[2] + lst[3]

            except IndexError:
                print("IndexError in File: ", filename, " Ln:", line_number)
                missed_lines +=1
            except UnicodeDecodeError:
                print("UnicodeDecodeError in File: ", filename, " Ln:", line_number)
                missed_lines += 1
            except csv.Error as e:
                print('file %s, line %d: %s' % (filename, line_number, e))
                missed_lines += 1
            except Exception:
                print("Exception in File: ", filename, " Ln:", line_number)
                missed_lines += 1
        total_missed_lines += missed_lines
        total_lines += line_number
        print("Total Missed lines: ", missed_lines, "/", line_number)

    return map_topic_description


# write csv/xls/xlsx file from dict object
def dict2csv(map_topic_description, output_xlsx_file):
    workbook = Workbook(output_xlsx_file)
    worksheet = workbook.add_worksheet()
    row = 0
    # Write Head fields
    worksheet.write(row, 0, "Topic")
    worksheet.write(row, 1, "Description")

    for topic in map_topic_description:
        row += 1
        worksheet.write(row, 0, topic)
        worksheet.write(row, 1, map_topic_description[topic])
    workbook.close()
# =======================================================


def tsv2xlsx(tsv_file, xlsx_file):
    # Create an XlsxWriter workbook object and add a worksheet.
    workbook = Workbook(xlsx_file)
    worksheet = workbook.add_worksheet()

    # Create a TSV file reader.
    tsv_reader = csv.reader(open(tsv_file, 'r'), delimiter='\t')

    # Read the row data from the TSV file and write it to the XLSX file.
    for row, data in enumerate(tsv_reader):
        worksheet.write_row(row, 0, data)

    # Close the XLSX file.
    workbook.close()

# read_csv_file("D:\\Projects\\DataTree\\data\\investopedia_data\\z.tsv")


def convert_tsv2xlsx(tsv_dir_path, xlsx_dir_path):
    for root, dirs, docs in os.walk(tsv_dir_path):
        # print root, dirs, files
        for doc in docs:
            if doc.endswith('.tsv'):
                csv_doc = os.path.join(root, doc)
                xlsx_doc = os.path.join(xlsx_dir_path, doc.split(".")[0]+".xlsx")
                print("Processing: ", csv_doc, xlsx_doc)
                try:
                    tsv2xlsx(csv_doc, xlsx_doc)
                except:
                    pass


def merge_xlsx_files(xlsx_dir_path, final_xlsx_file):
    outrow_idx = 0
    wkbk = xlwt.Workbook()
    outsheet = wkbk.add_sheet('Sheet1')

    for root, dirs, docs in os.walk(xlsx_dir_path):
        # print root, dirs, files
        for doc in docs:
            if doc.endswith('.xlsx'):
                xlsx_doc = os.path.join(root, doc)
                print("Processing: ", xlsx_doc)

                insheet = xlrd.open_workbook(xlsx_doc).sheets()[0]
                for row_idx in range(insheet.nrows):
                    for col_idx in range(insheet.ncols):
                        outsheet.write(outrow_idx, col_idx, insheet.cell_value(row_idx, col_idx))
                    outrow_idx += 1

    wkbk.save(final_xlsx_file)


# convert_tsv2xlsx(".\\data\\investopedia_data", ".\\data\\investopedia_data_xlsx")
# merge_xlsx_files(".\\data\\investopedia_data_xlsx", ".\\data\\investopedia_data_xlsx\\final_xlsx_file.xlsx")

# csv2dict(".\\data\\investopedia_data\\e.tsv")


# ==========================================Read & write xlsx file using xlrd, xlwt====================================
from xlrd import open_workbook

# returns list of rows (nested list)
def read_xlsx_xlrd(filename):
    wb = open_workbook(filename)
    values = []
    sheet = wb.sheet_by_index(0)
    #print 'Sheet:',s.name
    for row in range(0, sheet.nrows):
        col_value = []
        for col in range(0, sheet.ncols):
            col_value.append(sheet.cell(row, col).value)
        values.append(col_value)
    return values


# values : nested list of values(first line contains headers or heading)
def write_xlsx_xlwt(filename, values):
    book = xlwt.Workbook()
    sh = book.add_sheet("sheet1")

    r =0
    for row in values:
        c = 0
        for col in row:
            sh.write(r, c, col)
            c += 1
        r += 1

    book.save(filename)

# test
# values = [["col1", "col2", "col3"],[1,2,3,4],[2,3,4,5,6,7],[1,2,3]]
# write_xlsx_xlwt(".\\temp.xls", values)
# print(read_xlsx_xlrd(".\\temp.xls"))
# =======================================================================================